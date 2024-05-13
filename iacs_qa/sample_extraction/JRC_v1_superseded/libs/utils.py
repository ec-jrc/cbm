import os, shutil
import yaml as ym
import psycopg2 as ps
import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import formatting, styles
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

def load_yaml(file_name):

    with open(file_name, 'r') as stream:
        try:
            yaml_content = ym.safe_load(stream)
        except ym.YAMLError as exc:
            yaml_content = None
            print(f'ERROR loading yaml file {file_name}: {exc}')

    return yaml_content
    
def connect_db(db_params):
    return ps.connect(f"host={db_params['host']} dbname={db_params['db']} user={db_params['user']} port={db_params['port']} password={db_params['pwd']}")   

def create_folder(path):

    if not os.path.exists(path): 
        os.makedirs(path) 

def delete_folder(path):

    try:
        shutil.rmtree(path, ignore_errors=True)
    except:
        print(f"OSError to delete temporary folder: {path}")


def set_paths():

    pwd = Path(__file__).parent.parent.absolute()

    return dict(
        pwd = pwd,
        home = pwd.parent.absolute(),
    )

def save_csv(df, output_file):

    try:
        df.to_csv(output_file)
    except:
        df.to_csv(f'{output_file}.tmp.csv')
    finally:
        pass

def merge_lists_preserve_order(list1, list2):

    merged_list = list1.copy()  # Create a copy of the first list to preserve its original order
    
    for item in list2:
        if item not in merged_list:
            merged_list.append(item)
    
    return merged_list


def apply_xsl_style(ws):
    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'        


def create_new_xls(df, filename, sheetname, index=True, overwrite=True):      

    try:
    
        if overwrite and os.path.exists(filename):
            os.remove(filename)
        
        wb = Workbook()
        ws = wb.active    
        ws.title = sheetname  
        ws.sheet_view.showGridLines = False

        # ws = adjust_xls_column_width(ws)

        for r in dataframe_to_rows(df, index=index, header=True):
            ws.append(r)    

        wb.save(filename)  

    except:
        print(f"Error to create xls file (probably file is open): {filename}")

def add_sheet_to_xls(df, xls, sheetname, index=False):

    # try:

    wb = load_workbook(xls)    
    wb.create_sheet(sheetname)    
    ws = wb[sheetname]    
    ws.sheet_view.showGridLines = False
    ColumnDimension(ws, bestFit=True)
    
    # ws = columns_best_fit(ws)

    # ___ styles
    # apply_xsl_style(ws)  

    for r in dataframe_to_rows(df, index=index, header=True):
        ws.append(r)     
    wb.save(xls)

    # except:
    #     print(f"Error to create xls file (probably file is open): {xls}")
       


# def adjust_xls_column_width(ws):

#     dim_holder = DimensionHolder(worksheet=ws)

#     for col in range(ws.min_column, ws.max_column + 1):
#         dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=30)

#     ws.column_dimensions = dim_holder

#     return ws

# def columns_best_fit(ws):

#     column_widths = []
#     for row in data:
#         for i, cell in enumerate(row):
#             if len(column_widths) > i:
#                 if len(cell) > column_widths[i]:
#                     column_widths[i] = len(cell)
#             else:
#                 column_widths += [len(cell)]
        
#     for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
#         worksheet.column_dimensions[get_column_letter(i)].width = column_width

#     return ws

